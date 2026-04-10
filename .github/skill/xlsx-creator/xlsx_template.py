"""
XLSX Template — Microsoft-Branded Excel Dashboard Generator
Uses openpyxl to create professional workbooks with Dashboard + data tabs.

Usage:
    python3 xlsx_template.py

Requirements:
    pip install openpyxl

Customization:
    Replace the {{variables}} and sample data with real data.
    The agent should modify this template for each workbook request.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.formatting.rule import CellIsRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from datetime import date
import os

# === CONFIG ===
TITLE = "{{title}}"
AUTHOR = "{{author_name}}"
CLIENT = "{{client_name}}"
VERSION = "1.0.0"
DATE = date.today().isoformat()

# === MICROSOFT BRAND COLORS (hex without #) ===
BLUE = '0078D4'
RED = 'F25022'
GREEN = '7FBA00'
LIGHT_BLUE = '00A4EF'
YELLOW = 'FFB900'
DARK_GREEN = '107C10'
RED_ALERT = 'E81123'
TEXT_PRIMARY = '323130'
TEXT_SECONDARY = '605E5C'
ALT_ROW = 'F3F2F1'
BORDER_GRAY = 'D2D0CE'

# === STYLE PRESETS ===
header_fill = PatternFill(start_color=BLUE, fill_type='solid')
header_font = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
body_font = Font(name='Segoe UI', size=10, color=TEXT_PRIMARY)
title_font = Font(name='Segoe UI', size=18, bold=True, color=BLUE)
kpi_number_font = Font(name='Segoe UI', size=24, bold=True, color=BLUE)
kpi_label_font = Font(name='Segoe UI', size=9, color=TEXT_SECONDARY)
alt_fill = PatternFill(start_color=ALT_ROW, fill_type='solid')
white_fill = PatternFill(start_color='FFFFFF', fill_type='solid')
thin_border = Border(
    left=Side(style='thin', color=BORDER_GRAY),
    right=Side(style='thin', color=BORDER_GRAY),
    top=Side(style='thin', color=BORDER_GRAY),
    bottom=Side(style='thin', color=BORDER_GRAY),
)


def style_header_row(ws, row, max_col):
    """Apply Microsoft branding to a header row."""
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border


def style_data_rows(ws, start_row, end_row, max_col):
    """Apply alternating row colors and borders to data rows."""
    for row in range(start_row, end_row + 1):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = body_font
            cell.border = thin_border
            cell.fill = alt_fill if (row - start_row) % 2 == 0 else white_fill


def add_conditional_formatting(ws, col_letter, start_row, end_row):
    """Add status-based conditional formatting (Done=green, Blocked=red, In Progress=yellow)."""
    cell_range = f'{col_letter}{start_row}:{col_letter}{end_row}'

    ws.conditional_formatting.add(cell_range,
        CellIsRule(operator='equal', formula=['"Done"'],
                   fill=PatternFill(start_color='E8F5E9', fill_type='solid'),
                   font=Font(color=DARK_GREEN)))

    ws.conditional_formatting.add(cell_range,
        CellIsRule(operator='equal', formula=['"Blocked"'],
                   fill=PatternFill(start_color='FDEDED', fill_type='solid'),
                   font=Font(color=RED_ALERT)))

    ws.conditional_formatting.add(cell_range,
        CellIsRule(operator='equal', formula=['"In Progress"'],
                   fill=PatternFill(start_color='FFF3E0', fill_type='solid'),
                   font=Font(color='D83B01')))


def create_dashboard(wb):
    """Create the Dashboard tab with KPI cards and charts."""
    ws = wb.active
    ws.title = 'Dashboard'

    # Title bar
    ws.merge_cells('A1:H1')
    ws['A1'] = f'{TITLE} — {CLIENT}'
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A1'].fill = PatternFill(start_color=BLUE, fill_type='solid')
    ws['A1'].font = Font(name='Segoe UI', size=16, bold=True, color='FFFFFF')
    ws.row_dimensions[1].height = 40

    # Subtitle
    ws.merge_cells('A2:H2')
    ws['A2'] = f'v{VERSION} | {DATE} | {AUTHOR}'
    ws['A2'].font = Font(name='Segoe UI', size=10, color=TEXT_SECONDARY)
    ws['A2'].alignment = Alignment(horizontal='center')

    # KPI Cards (Row 4-5)
    kpis = [
        ('B4', '42', 'B5', 'Total Tasks'),
        ('D4', '78%', 'D5', 'Complete'),
        ('F4', '3', 'F5', 'Overdue'),
        ('H4', '5', 'H5', 'Blocked'),
    ]
    for val_cell, value, label_cell, label in kpis:
        ws[val_cell] = value
        ws[val_cell].font = kpi_number_font
        ws[val_cell].alignment = Alignment(horizontal='center')
        ws[label_cell] = label
        ws[label_cell].font = kpi_label_font
        ws[label_cell].alignment = Alignment(horizontal='center')

    # KPI formulas (connect to data tab)
    ws['B4'] = '=COUNTA(Tasks!B3:B100)'
    ws['D4'] = '=IFERROR(COUNTIF(Tasks!E3:E100,"Done")/COUNTA(Tasks!B3:B100),0)'
    ws['D4'].number_format = '0%'

    # Footer
    ws.merge_cells('A20:H20')
    ws['A20'] = f'Microsoft Confidential | Generated {DATE}'
    ws['A20'].font = Font(name='Segoe UI', size=8, italic=True, color=TEXT_SECONDARY)
    ws['A20'].alignment = Alignment(horizontal='center')

    # Column widths
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
        ws.column_dimensions[col].width = 15


def create_data_tab(wb, tab_name, headers, sample_data, status_col=None):
    """Create a branded data tab with headers, sample data, and formatting."""
    ws = wb.create_sheet(tab_name)

    # Tab title
    ws.merge_cells(f'A1:{get_column_letter(len(headers))}1')
    ws['A1'] = tab_name
    ws['A1'].font = Font(name='Segoe UI', size=14, bold=True, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color=BLUE, fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    # Column headers
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=2, column=col_idx, value=header)
    style_header_row(ws, 2, len(headers))

    # Sample data
    for row_idx, row_data in enumerate(sample_data, 3):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    style_data_rows(ws, 3, 2 + len(sample_data), len(headers))

    # Conditional formatting on status column
    if status_col:
        add_conditional_formatting(ws, status_col, 3, 100)

    # Data validation for status
    if status_col:
        status_val = DataValidation(
            type='list',
            formula1='"Not Started,In Progress,Blocked,Done"',
            allow_blank=True
        )
        status_val.prompt = 'Select status'
        ws.add_data_validation(status_val)
        status_val.add(f'{status_col}3:{status_col}100')

    # Freeze panes and auto-filter
    ws.freeze_panes = 'A3'
    ws.auto_filter.ref = f'A2:{get_column_letter(len(headers))}{2 + len(sample_data)}'

    # Column widths
    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = max(15, len(headers[col_idx - 1]) + 5)

    return ws


def create_workbook():
    """Generate the complete branded workbook."""
    wb = openpyxl.Workbook()

    # === DASHBOARD ===
    create_dashboard(wb)

    # === TASKS TAB ===
    create_data_tab(wb, 'Tasks',
        headers=['#', 'Task', 'Owner', 'Due Date', 'Status', 'Priority'],
        sample_data=[
            [1, 'Setup project scaffold', 'Team Lead', '2026-03-10', 'Done', 'High'],
            [2, 'Design architecture', 'Architect', '2026-03-15', 'In Progress', 'High'],
            [3, 'Implement API', 'Developer', '2026-03-20', 'Not Started', 'Medium'],
            [4, 'Security review', 'Security', '2026-03-25', 'Not Started', 'High'],
            [5, 'User testing', 'QA', '2026-03-30', 'Not Started', 'Medium'],
        ],
        status_col='E'
    )

    # === MILESTONES TAB ===
    create_data_tab(wb, 'Milestones',
        headers=['#', 'Milestone', 'Target Date', 'Status', 'Notes'],
        sample_data=[
            [1, 'Phase 1 Complete', '2026-03-31', 'In Progress', 'On track'],
            [2, 'Phase 2 Kickoff', '2026-04-01', 'Not Started', 'Pending Phase 1'],
            [3, 'Final Delivery', '2026-06-30', 'Not Started', ''],
        ],
        status_col='D'
    )

    # === ADD CHART TO DASHBOARD ===
    ws_tasks = wb['Tasks']
    ws_dashboard = wb['Dashboard']

    chart = BarChart()
    chart.type = 'col'
    chart.title = 'Task Status Distribution'
    chart.style = 10
    chart.width = 20
    chart.height = 12
    # Note: In production, the agent calculates chart data from the tasks
    ws_dashboard.add_chart(chart, 'A8')

    # === SAVE ===
    os.makedirs('output/xlsx', exist_ok=True)
    filename = f"{TITLE.replace(' ', '_')}_v{VERSION}_{DATE}.xlsx"
    filepath = f"output/xlsx/{filename}"
    wb.save(filepath)
    print(f"Created: {filepath}")


if __name__ == '__main__':
    create_workbook()
